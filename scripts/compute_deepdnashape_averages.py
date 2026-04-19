#!/usr/bin/env python3
"""
Compute average DNA shape features for Sequence_9_Long using Deep DNAshape.

Expected columns in input Excel:
- Sequence_9_Long

Adds/updates columns:
- Average_MGW
- Average_ProT
- Average_HelT
- Average_Roll
- Average_EP
"""

from __future__ import annotations

import argparse
import importlib
import math
import platform
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Any

from openpyxl import load_workbook


SEQ_COL = "Sequence_9_Long"
FEATURE_TO_COL = {
    "MGW": "Average_MGW",
    "ProT": "Average_ProT",
    "HelT": "Average_HelT",
    "Roll": "Average_Roll",
    "EP": "Average_EP",
}
VALID_SEQ = re.compile(r"^[ACGTNacgtn]+$")


def run(cmd: List[str], check: bool = True) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, text=True, capture_output=True, check=check)


def clean_seq(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    if not VALID_SEQ.match(s):
        return None
    return s


def install_deepdnashape(python_exe: str) -> None:
    # Install a dependency set compatible with deepDNAshape model weight loading.
    # deepDNAshape uses legacy Keras2-style load_weights paths.
    is_macos = platform.system().lower() == "darwin"
    is_arm = platform.machine().lower() in {"arm64", "aarch64"}

    if is_macos and is_arm:
        pre_cmd = [
            python_exe,
            "-m",
            "pip",
            "install",
            "numpy<2",
            "keras<3",
            "tensorflow-macos==2.15.0",
        ]
    else:
        pre_cmd = [
            python_exe,
            "-m",
            "pip",
            "install",
            "numpy<2",
            "keras<3",
            "tensorflow==2.15.1",
        ]

    pre = run(pre_cmd, check=False)
    if pre.returncode != 0:
        raise RuntimeError(
            "Failed to install compatible TensorFlow/Keras dependencies.\n"
            f"STDOUT:\n{pre.stdout}\nSTDERR:\n{pre.stderr}"
        )

    # Official code location per Nature Communications paper.
    cp = run([python_exe, "-m", "pip", "install", "git+https://github.com/JinsenLi/deepDNAshape.git"], check=False)
    if cp.returncode != 0:
        raise RuntimeError(
            "Failed to install deepDNAshape from GitHub.\n"
            f"Command: {python_exe} -m pip install git+https://github.com/JinsenLi/deepDNAshape.git\n"
            f"STDOUT:\n{cp.stdout}\nSTDERR:\n{cp.stderr}"
        )


def find_deepdnashape_executable() -> Optional[str]:
    exe = shutil.which("deepDNAshape")
    if exe:
        return exe
    # Try current Python environment's bin/Scripts folder.
    base = Path(sys.executable).resolve().parent
    candidates = [base / "deepDNAshape", base / "Scripts" / "deepDNAshape.exe"]
    for c in candidates:
        if c.exists():
            return str(c)
    return None


def resolve_deepdnashape_cmd() -> Optional[List[str]]:
    exe = find_deepdnashape_executable()
    if exe:
        return [exe]
    # Fallback when console entrypoint is missing.
    p = subprocess.run([sys.executable, "-m", "deepDNAshape", "--help"], text=True, capture_output=True)
    if p.returncode == 0:
        return [sys.executable, "-m", "deepDNAshape"]
    return None


def load_deepdnashape_predictor() -> Optional[Any]:
    """
    Load deepDNAshape predictor object directly from Python package when CLI is unavailable.
    """
    try:
        mod = importlib.import_module("deepDNAshape.predictor")
    except Exception:
        return None

    # Upstream has used lowercase class naming in some versions.
    for cls_name in ("predictor", "Predictor"):
        cls = getattr(mod, cls_name, None)
        if isinstance(cls, type):
            try:
                return cls()
            except Exception:
                continue
    return None


def deepdnashape_supported_features(cmd_prefix: Sequence[str]) -> List[str]:
    p = subprocess.run([*cmd_prefix, "--help"], text=True, capture_output=True)
    text = (p.stdout or "") + "\n" + (p.stderr or "")
    # Parse "... other options: Shear, Stretch, ..., HelT]."
    m = re.search(r"other options:\s*([^\]]+)\]", text, flags=re.IGNORECASE)
    features = ["MGW"]
    if m:
        extras = [x.strip() for x in m.group(1).split(",") if x.strip()]
        features.extend(extras)
    return sorted(set(features))


def parse_prediction_line(line: str) -> Optional[float]:
    vals = [x for x in line.strip().replace(",", " ").split() if x]
    if not vals:
        return None
    nums: List[float] = []
    for v in vals:
        try:
            nums.append(float(v))
        except ValueError:
            continue
    if not nums:
        return None
    return float(sum(nums) / len(nums))


def predict_feature_averages(
    cmd_prefix: Sequence[str],
    sequences: List[str],
    feature: str,
    layer: int,
) -> List[Optional[float]]:
    with tempfile.TemporaryDirectory(prefix="deepdnashape_") as td:
        td_path = Path(td)
        in_txt = td_path / "seqs.txt"
        out_txt = td_path / "pred.txt"
        in_txt.write_text("\n".join(sequences) + "\n", encoding="utf-8")

        p = run(
            [*cmd_prefix, "--file", str(in_txt), "--feature", feature, "--layer", str(layer), "--output", str(out_txt)],
            check=False,
        )
        if p.returncode != 0:
            msg = (p.stdout or "") + "\n" + (p.stderr or "")
            if "Keras 3 only supports" in msg:
                raise RuntimeError(
                    "deepDNAshape failed due to Keras 3 incompatibility.\n"
                    "Fix by pinning legacy stack in this venv:\n"
                    "pip uninstall -y keras tensorflow tensorflow-macos\n"
                    "pip install 'numpy<2' 'keras<3' 'tensorflow-macos==2.15.0'\n"
                    "pip install --force-reinstall git+https://github.com/JinsenLi/deepDNAshape.git"
                )
            raise RuntimeError(
                f"deepDNAshape failed for feature {feature}.\nSTDOUT:\n{p.stdout}\nSTDERR:\n{p.stderr}"
            )
        if not out_txt.exists():
            raise RuntimeError(f"deepDNAshape did not create output file for feature {feature}.")

        lines = [ln.rstrip("\n") for ln in out_txt.read_text(encoding="utf-8").splitlines()]
        if len(lines) != len(sequences):
            raise RuntimeError(
                f"Output line count mismatch for {feature}: got {len(lines)}, expected {len(sequences)}."
            )
        return [parse_prediction_line(ln) for ln in lines]


def predict_feature_averages_api(
    predictor_obj: Any,
    sequences: List[str],
    feature: str,
    layer: int,
) -> List[Optional[float]]:
    out: List[Optional[float]] = []
    for seq in sequences:
        try:
            pred = predictor_obj.predict(feature, seq, layer)
        except Exception as e:
            msg = str(e)
            if "Keras 3 only supports" in msg:
                raise RuntimeError(
                    "deepDNAshape failed due to Keras 3 incompatibility.\n"
                    "Fix by pinning legacy stack in this venv:\n"
                    "pip uninstall -y keras tensorflow tensorflow-macos\n"
                    "pip install 'numpy<2' 'keras<3' 'tensorflow-macos==2.15.0'\n"
                    "pip install --force-reinstall git+https://github.com/JinsenLi/deepDNAshape.git"
                ) from e
            raise RuntimeError(f"deepDNAshape API prediction failed for feature {feature}: {e}") from e

        if pred is None:
            out.append(None)
            continue

        nums: List[float] = []
        for v in pred:
            try:
                nums.append(float(v))
            except Exception:
                continue
        out.append(float(sum(nums) / len(nums)) if nums else None)
    return out


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Compute Deep DNAshape average features from 9-mer sequences.")
    p.add_argument("--input-xlsx", required=True, help="Input workbook path.")
    p.add_argument("--output-xlsx", required=True, help="Output workbook path.")
    p.add_argument("--sheet", default="Hotspots", help="Sheet name (default: Hotspots).")
    p.add_argument("--layer", type=int, default=4, help="Deep DNAshape layer (2/3/4/5 recommended for web parity).")
    p.add_argument("--install-deepdnashape", action="store_true", help="Install deepDNAshape into current env.")
    p.add_argument(
        "--strict-ep",
        action="store_true",
        help="Fail if EP is not supported by installed deepDNAshape build.",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if sys.version_info >= (3, 13):
        raise RuntimeError(
            f"Python {sys.version.split()[0]} is too new for typical TensorFlow/deepDNAshape dependency wheels. "
            "Use Python 3.10, 3.11, or 3.12 in a fresh virtual environment."
        )

    if args.install_deepdnashape:
        install_deepdnashape(sys.executable)

    deep_cmd = resolve_deepdnashape_cmd()
    predictor_obj = None
    if not deep_cmd:
        predictor_obj = load_deepdnashape_predictor()
    if not deep_cmd and predictor_obj is None:
        raise RuntimeError(
            "deepDNAshape is installed but neither CLI command nor Python predictor API could be loaded. "
            "Try reinstalling: pip install --force-reinstall --no-deps git+https://github.com/JinsenLi/deepDNAshape.git"
        )

    wb = load_workbook(args.input_xlsx)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found in workbook.")
    ws = wb[args.sheet]

    # Header row is expected in row 1.
    header_cells = [cell.value for cell in ws[1]]
    header_map: Dict[str, int] = {}
    for idx, name in enumerate(header_cells, start=1):
        if isinstance(name, str) and name.strip():
            header_map[name.strip()] = idx

    if SEQ_COL not in header_map:
        raise ValueError(f"Missing column: {SEQ_COL}")

    max_col = ws.max_column
    for out_col in FEATURE_TO_COL.values():
        if out_col not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=out_col)
            header_map[out_col] = max_col

    seq_col_idx = header_map[SEQ_COL]
    data_rows: List[int] = []
    seqs: List[str] = []
    for row_idx in range(2, ws.max_row + 1):
        data_rows.append(row_idx)
        raw = ws.cell(row=row_idx, column=seq_col_idx).value
        cleaned = clean_seq(raw)
        seqs.append(cleaned if cleaned is not None else "")

    valid_mask = [bool(s) for s in seqs]

    # Keep original row alignment; predict only valid rows.
    valid_seqs = [s for s in seqs if s]

    feature_values: Dict[str, List[Optional[float]]] = {f: [None] * len(seqs) for f in FEATURE_TO_COL}

    if deep_cmd:
        supported_features = set(deepdnashape_supported_features(deep_cmd))
    else:
        # Conservative hardcoded set from deepDNAshape help text.
        supported_features = {
            "MGW",
            "Shear",
            "Stretch",
            "Stagger",
            "Buckle",
            "ProT",
            "Opening",
            "Shift",
            "Slide",
            "Rise",
            "Tilt",
            "Roll",
            "HelT",
        }

    for feature in FEATURE_TO_COL:
        supported = feature in supported_features
        if not supported:
            if feature == "EP" and not args.strict_ep:
                print(
                    "Warning: Installed deepDNAshape build does not support EP; "
                    "Average_EP will be left empty."
                )
                continue
            raise RuntimeError(
                f"Feature '{feature}' is not supported by this deepDNAshape installation."
            )

        if deep_cmd:
            averages_valid = predict_feature_averages(deep_cmd, valid_seqs, feature=feature, layer=args.layer)
        else:
            averages_valid = predict_feature_averages_api(predictor_obj, valid_seqs, feature=feature, layer=args.layer)
        j = 0
        for i, is_valid in enumerate(valid_mask):
            if is_valid:
                feature_values[feature][i] = averages_valid[j]
                j += 1

    for feature, col in FEATURE_TO_COL.items():
        col_idx = header_map[col]
        values = feature_values[feature]
        for i, row_idx in enumerate(data_rows):
            ws.cell(row=row_idx, column=col_idx, value=values[i])

    wb.save(args.output_xlsx)

    print(f"Wrote: {args.output_xlsx}")
    print(f"Rows: {len(seqs)}")
    if deep_cmd:
        print(f"DeepDNAshape command: {' '.join(deep_cmd)}")
    else:
        print("DeepDNAshape mode: python API (deepDNAshape.predictor)")


if __name__ == "__main__":
    main()
