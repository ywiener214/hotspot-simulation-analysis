#!/usr/bin/env python3
"""
Generate GC-matched controls, recalc MGW/ProT via deepDNAshape, and run KS tests.

Inputs:
- Excel with hotspots (Sequence_9_Long, Sequence_15_Long, Gene)
- FASTA directory with <Gene>.fasta files

Outputs:
- New Excel with:
  - Hotspot sheet updated with Average_MGW/ProT recalculated
  - New sheet "Controls" with GC-matched controls + MGW/ProT
- KS test summary printed to stdout

GC content:
  GC_count = number of G/C bases in Sequence_9_Long (0..9)
  GC_fraction = GC_count / 9
Matching:
  Controls are sampled from nearby non-overlapping candidates, then matched to
  hotspots by exact GC_count. If insufficient, optional relaxed matching can be enabled.
"""

from __future__ import annotations

import argparse
import math
import random
import re
import shutil
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple, Any

import importlib

from openpyxl import load_workbook, Workbook


DNA = set("ACGTN")
SEQ9_COL = "Sequence_9_Long"
SEQ15_COL = "Sequence_15_Long"
GENE_COL = "Gene"
OUT_CTRL_SHEET = "Controls"
RING_FEATURES = ["MGW", "ProT"]


def clean_seq(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    s = "".join(ch for ch in s if ch in DNA)
    return s or None


def gc_count(seq: str) -> int:
    return sum(1 for ch in seq if ch in ("G", "C"))


def gc_fraction(seq: str) -> float:
    return gc_count(seq) / len(seq) if seq else 0.0


def read_fasta_sequence(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Missing FASTA file: {path}")
    chunks: List[str] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith(">"):
                continue
            chunks.append(line.upper())
    if not chunks:
        raise ValueError(f"No sequence content in FASTA: {path}")
    return "".join(chunks)


def matches_at(target: str, query: str, start: int) -> bool:
    for i, q in enumerate(query):
        t = target[start + i]
        if q != "N" and t != "N" and q != t:
            return False
    return True


def find_all_matches(target: str, query: str) -> List[int]:
    if not query or len(query) > len(target):
        return []
    out: List[int] = []
    max_start = len(target) - len(query)
    for s in range(max_start + 1):
        if matches_at(target, query, s):
            out.append(s)
    return out


def intervals_overlap(a_start: int, a_end: int, b_start: int, b_end: int, min_gap: int) -> bool:
    return not (a_end + min_gap < b_start or b_end + min_gap < a_start)


def run(cmd: List[str], check: bool = True) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, text=True, capture_output=True, check=check)


def find_deepdnashape_cmd() -> Optional[List[str]]:
    exe = shutil.which("deepDNAshape")
    if exe:
        return [exe]
    base = Path(sys.executable).resolve().parent
    candidates = [base / "deepDNAshape", base / "Scripts" / "deepDNAshape.exe"]
    for c in candidates:
        if c.exists():
            return [str(c)]
    # Fallback: python -m deepDNAshape
    p = subprocess.run([sys.executable, "-m", "deepDNAshape", "--help"], text=True, capture_output=True)
    if p.returncode == 0:
        return [sys.executable, "-m", "deepDNAshape"]
    return None


def load_deepdnashape_predictor() -> Optional[Any]:
    try:
        mod = importlib.import_module("deepDNAshape.predictor")
    except Exception:
        return None
    for cls_name in ("predictor", "Predictor"):
        cls = getattr(mod, cls_name, None)
        if isinstance(cls, type):
            try:
                return cls()
            except Exception:
                continue
    return None


def predict_feature_averages(cmd_prefix: Sequence[str], sequences: List[str], feature: str, layer: int) -> List[Optional[float]]:
    with tempfile.TemporaryDirectory(prefix="deepdnashape_") as td:
        td_path = Path(td)
        in_txt = td_path / "seqs.txt"
        out_txt = td_path / "pred.txt"
        in_txt.write_text("\n".join(sequences) + "\n", encoding="utf-8")

        p = run([*cmd_prefix, "--file", str(in_txt), "--feature", feature, "--layer", str(layer), "--output", str(out_txt)], check=False)
        if p.returncode != 0:
            msg = (p.stdout or "") + "\n" + (p.stderr or "")
            raise RuntimeError(f"deepDNAshape failed for {feature}.\n{msg}")
        if not out_txt.exists():
            raise RuntimeError(f"deepDNAshape did not create output for {feature}.")

        lines = out_txt.read_text(encoding="utf-8").splitlines()
        if len(lines) != len(sequences):
            raise RuntimeError(f"Output line count mismatch for {feature}: {len(lines)} vs {len(sequences)}.")

        out: List[Optional[float]] = []
        for ln in lines:
            parts = [x for x in ln.strip().replace(",", " ").split() if x]
            vals: List[float] = []
            for v in parts:
                try:
                    vals.append(float(v))
                except ValueError:
                    continue
            out.append(sum(vals) / len(vals) if vals else None)
        return out


def predict_feature_averages_api(predictor_obj: Any, sequences: List[str], feature: str, layer: int) -> List[Optional[float]]:
    out: List[Optional[float]] = []
    for seq in sequences:
        try:
            pred = predictor_obj.predict(feature, seq, layer)
        except Exception as e:
            msg = str(e)
            raise RuntimeError(f"deepDNAshape API failed for {feature}: {msg}") from e
        if pred is None:
            out.append(None)
            continue
        vals: List[float] = []
        for v in pred:
            try:
                vals.append(float(v))
            except Exception:
                continue
        out.append(sum(vals) / len(vals) if vals else None)
    return out


def ks_test_2samp(x: List[float], y: List[float]) -> Tuple[float, float]:
    # Two-sample KS test (asymptotic p-value).
    x = sorted([v for v in x if v is not None])
    y = sorted([v for v in y if v is not None])
    n1, n2 = len(x), len(y)
    if n1 == 0 or n2 == 0:
        return float("nan"), float("nan")
    i = j = 0
    d = 0.0
    while i < n1 and j < n2:
        if x[i] <= y[j]:
            i += 1
        else:
            j += 1
        d = max(d, abs(i / n1 - j / n2))
    # asymptotic p-value
    en = math.sqrt(n1 * n2 / (n1 + n2))
    lam = (en + 0.12 + 0.11 / en) * d
    # Kolmogorov distribution approximation
    p = 2.0 * sum(((-1) ** (k - 1)) * math.exp(-2 * (lam ** 2) * (k ** 2)) for k in range(1, 100))
    p = max(0.0, min(1.0, p))
    return d, p


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--input-xlsx", required=True, help="Hotspot workbook path")
    p.add_argument("--output-xlsx", required=True, help="Output workbook path (combined hotspots + controls)")
    p.add_argument("--sheet", default=None, help="Hotspot sheet name (default: first sheet)")
    p.add_argument("--fasta-dir", required=True, help="Directory of <Gene>.fasta files")
    p.add_argument("--target-controls", type=int, default=5000)
    p.add_argument("--nearby-max-distance", type=int, default=500)
    p.add_argument("--min-gap", type=int, default=0)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--layer", type=int, default=4)
    p.add_argument("--relax-gc", action="store_true", help="Allow fallback GC matching +/-1 if exact unavailable")
    p.add_argument(
        "--controls-only-xlsx",
        default=None,
        help="If set, write combined Hotspots/Controls workbook here before KS test.",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    rng = random.Random(args.seed)

    wb = load_workbook(args.input_xlsx)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    col_map: Dict[str, int] = {}
    for idx, name in enumerate(header, start=1):
        if isinstance(name, str) and name.strip():
            col_map[name.strip()] = idx

    for required in (GENE_COL, SEQ9_COL, SEQ15_COL):
        if required not in col_map:
            raise ValueError(f"Missing column: {required}")

    rows: List[Dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        gene = ws.cell(r, col_map[GENE_COL]).value
        seq9 = clean_seq(ws.cell(r, col_map[SEQ9_COL]).value)
        seq15 = clean_seq(ws.cell(r, col_map[SEQ15_COL]).value)
        if gene and seq9 and seq15:
            rows.append({"row": r, "gene": str(gene).strip(), "seq9": seq9, "seq15": seq15})

    genes = sorted({r["gene"] for r in rows})
    fasta_dir = Path(args.fasta_dir)
    gene_seq: Dict[str, str] = {g: read_fasta_sequence(fasta_dir / f"{g}.fasta") for g in genes}

    # Map hotspot centers
    hotspot_centers_by_gene: Dict[str, List[int]] = defaultdict(list)
    hotspot_intervals_by_gene: Dict[str, List[Tuple[int, int]]] = defaultdict(list)
    hotspot_seq9_set = {r["seq9"] for r in rows}
    hotspot_seq15_set = {r["seq15"] for r in rows}

    for r in rows:
        gseq = gene_seq[r["gene"]]
        centers: List[int] = []
        if r["seq15"] and len(r["seq15"]) == 15:
            for s in find_all_matches(gseq, r["seq15"]):
                centers.append(s + 7)
        if not centers and r["seq9"] and len(r["seq9"]) == 9:
            for s in find_all_matches(gseq, r["seq9"]):
                centers.append(s + 4)
        for c in centers:
            hotspot_centers_by_gene[r["gene"]].append(c)
            hs_start, hs_end = max(0, c - 7), min(len(gseq) - 1, c + 7)
            hotspot_intervals_by_gene[r["gene"]].append((hs_start, hs_end))

    # Build candidate controls per gene/GC count
    candidates_by_gene_gc: Dict[str, Dict[int, List[int]]] = defaultdict(lambda: defaultdict(list))
    for g in genes:
        gseq = gene_seq[g]
        centers = hotspot_centers_by_gene.get(g, [])
        protected = hotspot_intervals_by_gene.get(g, [])
        if not centers:
            continue
        for hc in centers:
            low = max(7, hc - args.nearby_max_distance)
            high = min(len(gseq) - 8, hc + args.nearby_max_distance)
            for c in range(low, high + 1):
                c15_start, c15_end = c - 7, c + 7
                c9_start, c9_end = c - 4, c + 4
                blocked = False
                for hs, he in protected:
                    if intervals_overlap(c15_start, c15_end, hs, he, args.min_gap):
                        blocked = True
                        break
                if blocked:
                    continue
                seq15 = gseq[c15_start : c15_end + 1]
                seq9 = gseq[c9_start : c9_end + 1]
                if len(seq15) != 15 or len(seq9) != 9:
                    continue
                if seq9 in hotspot_seq9_set or seq15 in hotspot_seq15_set:
                    continue
                gc = gc_count(seq9)
                candidates_by_gene_gc[g][gc].append(c)

    # GC-matched sampling: one control per hotspot row
    controls: List[Dict[str, object]] = []
    deficit = Counter()
    for r in rows:
        g = r["gene"]
        seq9 = r["seq9"]
        gc = gc_count(seq9)
        pool = candidates_by_gene_gc[g].get(gc, [])
        if not pool and args.relax_gc:
            # try nearest GC counts
            for delta in (1, 2, 3):
                for gc2 in (gc - delta, gc + delta):
                    if 0 <= gc2 <= 9 and candidates_by_gene_gc[g].get(gc2):
                        pool = candidates_by_gene_gc[g][gc2]
                        break
                if pool:
                    break
        if not pool:
            deficit[(g, gc)] += 1
            continue
        c = pool.pop(rng.randrange(len(pool)))
        gseq = gene_seq[g]
        seq15 = gseq[c - 7 : c + 8]
        seq9c = gseq[c - 4 : c + 5]
        controls.append(
            {
                "gene": g,
                "seq9": seq9c,
                "seq15": seq15,
                "gc_count": gc_count(seq9c),
                "gc_fraction": gc_fraction(seq9c),
            }
        )

    # Recalculate MGW/ProT for hotspots and controls
    cmd = find_deepdnashape_cmd()
    predictor_obj = None
    if not cmd:
        predictor_obj = load_deepdnashape_predictor()
    if not cmd and predictor_obj is None:
        raise RuntimeError("deepDNAshape not available in this environment.")

    hotspot_seq9 = [r["seq9"] for r in rows]
    control_seq9 = [r["seq9"] for r in controls]

    if cmd:
        mgw_hot = predict_feature_averages(cmd, hotspot_seq9, "MGW", args.layer)
        prot_hot = predict_feature_averages(cmd, hotspot_seq9, "ProT", args.layer)
        mgw_ctl = predict_feature_averages(cmd, control_seq9, "MGW", args.layer)
        prot_ctl = predict_feature_averages(cmd, control_seq9, "ProT", args.layer)
    else:
        mgw_hot = predict_feature_averages_api(predictor_obj, hotspot_seq9, "MGW", args.layer)
        prot_hot = predict_feature_averages_api(predictor_obj, hotspot_seq9, "ProT", args.layer)
        mgw_ctl = predict_feature_averages_api(predictor_obj, control_seq9, "MGW", args.layer)
        prot_ctl = predict_feature_averages_api(predictor_obj, control_seq9, "ProT", args.layer)

    # Write back to hotspot sheet (overwrite Average_MGW / Average_ProT)
    for name in ("Average_MGW", "Average_ProT"):
        if name not in col_map:
            col_map[name] = ws.max_column + 1
            ws.cell(1, col_map[name], value=name)

    for i, r in enumerate(rows):
        ws.cell(r["row"], col_map["Average_MGW"], value=mgw_hot[i])
        ws.cell(r["row"], col_map["Average_ProT"], value=prot_hot[i])

    # Create controls sheet
    if OUT_CTRL_SHEET in wb.sheetnames:
        del wb[OUT_CTRL_SHEET]
    ws_ctrl = wb.create_sheet(OUT_CTRL_SHEET)
    headers = ["Gene", "Sequence_15_Long", "Sequence_9_Long", "GC_Count", "GC_Fraction", "Average_MGW", "Average_ProT"]
    ws_ctrl.append(headers)
    for i, c in enumerate(controls):
        ws_ctrl.append(
            [
                c["gene"],
                c["seq15"],
                c["seq9"],
                c["gc_count"],
                c["gc_fraction"],
                mgw_ctl[i] if i < len(mgw_ctl) else None,
                prot_ctl[i] if i < len(prot_ctl) else None,
            ]
        )

    # Write combined workbook (Hotspots updated + Controls sheet).
    wb.save(args.controls_only_xlsx or args.output_xlsx)

    # KS tests (after workbook is written)
    d_mgw, p_mgw = ks_test_2samp([v for v in mgw_hot if v is not None], [v for v in mgw_ctl if v is not None])
    d_prot, p_prot = ks_test_2samp([v for v in prot_hot if v is not None], [v for v in prot_ctl if v is not None])

    print("GC content definition: GC_count = count(G/C) in 9-mer; GC_fraction = GC_count/9.")
    print("Matching: exact GC_count within same-gene candidate pool (nearby, non-overlapping).")
    if deficit:
        print(f"Unmatched hotspots (by gene, gc): {sum(deficit.values())}")
    print(f"Controls generated: {len(controls)}")
    print(f"KS test MGW: D={d_mgw:.4f}, p={p_mgw:.4g}")
    print(f"KS test ProT: D={d_prot:.4f}, p={p_prot:.4g}")

    # If controls_only_xlsx was used, also write a KS summary file path hint.
    if args.controls_only_xlsx and args.controls_only_xlsx != args.output_xlsx:
        print(f"Combined workbook written to: {args.controls_only_xlsx}")
        print(f"Final output workbook written to: {args.output_xlsx}")


if __name__ == "__main__":
    main()
