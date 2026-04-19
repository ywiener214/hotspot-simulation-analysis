#!/usr/bin/env python3
"""
Generate controls (default 10,000) and compute MGW/ProT for both hotspots and controls.

Inputs:
- Hotspot workbook with Gene, Sequence_9_Long, Sequence_15_Long
- FASTA directory with <Gene>.fasta files

Output:
- Excel with two sheets:
  Hotspots (copied from input with recalculated Average_MGW / Average_ProT)
  Controls (generated with Average_MGW / Average_ProT)
"""

from __future__ import annotations

import argparse
import math
import random
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple, Any
import importlib

from openpyxl import Workbook, load_workbook


DNA = set("ACGTN")
SEQ9_COL = "Sequence_9_Long"
SEQ15_COL = "Sequence_15_Long"
GENE_COL = "Gene"


def clean_seq(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    s = "".join(ch for ch in s if ch in DNA)
    return s or None


def read_fasta_sequence(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Missing FASTA file: {path}")
    chunks: List[str] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith(">"):
                continue
            chunks.append(line.upper())
    if not chunks:
        raise ValueError(f"No sequence content in FASTA: {path}")
    return "".join(chunks)


def matches_at(target: str, query: str, start: int) -> bool:
    for i, q in enumerate(query):
        t = target[start + i]
        if q != "N" and t != "N" and q != t:
            return False
    return True


def find_all_matches(target: str, query: str) -> List[int]:
    if not query or len(query) > len(target):
        return []
    out: List[int] = []
    max_start = len(target) - len(query)
    for s in range(max_start + 1):
        if matches_at(target, query, s):
            out.append(s)
    return out


def intervals_overlap(a_start: int, a_end: int, b_start: int, b_end: int, min_gap: int) -> bool:
    return not (a_end + min_gap < b_start or b_end + min_gap < a_start)


def run(cmd: List[str], check: bool = True) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, text=True, capture_output=True, check=check)


def find_deepdnashape_cmd() -> Optional[List[str]]:
    exe = shutil.which("deepDNAshape")
    if exe:
        return [exe]
    base = Path(sys.executable).resolve().parent
    candidates = [base / "deepDNAshape", base / "Scripts" / "deepDNAshape.exe"]
    for c in candidates:
        if c.exists():
            return [str(c)]
    p = subprocess.run([sys.executable, "-m", "deepDNAshape", "--help"], text=True, capture_output=True)
    if p.returncode == 0:
        return [sys.executable, "-m", "deepDNAshape"]
    return None


def load_deepdnashape_predictor() -> Optional[Any]:
    try:
        mod = importlib.import_module("deepDNAshape.predictor")
    except Exception:
        return None
    for cls_name in ("predictor", "Predictor"):
        cls = getattr(mod, cls_name, None)
        if isinstance(cls, type):
            try:
                return cls()
            except Exception:
                continue
    return None


def predict_feature_averages(cmd_prefix: Sequence[str], sequences: List[str], feature: str, layer: int) -> List[Optional[float]]:
    with tempfile.TemporaryDirectory(prefix="deepdnashape_") as td:
        td_path = Path(td)
        in_txt = td_path / "seqs.txt"
        out_txt = td_path / "pred.txt"
        in_txt.write_text("\n".join(sequences) + "\n", encoding="utf-8")

        p = run([*cmd_prefix, "--file", str(in_txt), "--feature", feature, "--layer", str(layer), "--output", str(out_txt)], check=False)
        if p.returncode != 0:
            msg = (p.stdout or "") + "\n" + (p.stderr or "")
            raise RuntimeError(f"deepDNAshape failed for {feature}.\n{msg}")
        if not out_txt.exists():
            raise RuntimeError(f"deepDNAshape did not create output for {feature}.")

        lines = out_txt.read_text(encoding="utf-8").splitlines()
        if len(lines) != len(sequences):
            raise RuntimeError(f"Output line count mismatch for {feature}: {len(lines)} vs {len(sequences)}.")

        out: List[Optional[float]] = []
        for ln in lines:
            parts = [x for x in ln.strip().replace(",", " ").split() if x]
            vals: List[float] = []
            for v in parts:
                try:
                    vals.append(float(v))
                except ValueError:
                    continue
            out.append(sum(vals) / len(vals) if vals else None)
        return out


def predict_feature_averages_api(predictor_obj: Any, sequences: List[str], feature: str, layer: int) -> List[Optional[float]]:
    out: List[Optional[float]] = []
    for seq in sequences:
        pred = predictor_obj.predict(feature, seq, layer)
        if pred is None:
            out.append(None)
            continue
        vals: List[float] = []
        for v in pred:
            try:
                vals.append(float(v))
            except Exception:
                continue
        out.append(sum(vals) / len(vals) if vals else None)
    return out


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--input-xlsx", required=True, help="Hotspot workbook path")
    p.add_argument("--output-xlsx", required=True, help="Output workbook path")
    p.add_argument("--sheet", default=None, help="Hotspot sheet name (default: first sheet)")
    p.add_argument("--fasta-dir", required=True, help="Directory of <Gene>.fasta files")
    p.add_argument("--target-controls", type=int, default=10000)
    p.add_argument("--nearby-max-distance", type=int, default=500)
    p.add_argument("--min-gap", type=int, default=0)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--layer", type=int, default=4)
    return p.parse_args()


def main() -> None:
    args = parse_args()
    rng = random.Random(args.seed)

    wb = load_workbook(args.input_xlsx)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    col_map: Dict[str, int] = {}
    for idx, name in enumerate(header, start=1):
        if isinstance(name, str) and name.strip():
            col_map[name.strip()] = idx

    for required in (GENE_COL, SEQ9_COL, SEQ15_COL):
        if required not in col_map:
            raise ValueError(f"Missing column: {required}")

    hotspot_rows: List[Dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        gene = ws.cell(r, col_map[GENE_COL]).value
        seq9 = clean_seq(ws.cell(r, col_map[SEQ9_COL]).value)
        seq15 = clean_seq(ws.cell(r, col_map[SEQ15_COL]).value)
        if gene and seq9 and seq15:
            hotspot_rows.append({"gene": str(gene).strip(), "seq9": seq9, "seq15": seq15})

    genes = sorted({r["gene"] for r in hotspot_rows})
    fasta_dir = Path(args.fasta_dir)
    gene_seq: Dict[str, str] = {g: read_fasta_sequence(fasta_dir / f"{g}.fasta") for g in genes}

    hotspot_centers_by_gene: Dict[str, List[int]] = defaultdict(list)
    hotspot_intervals_by_gene: Dict[str, List[Tuple[int, int]]] = defaultdict(list)
    hotspot_seq9_set = {r["seq9"] for r in hotspot_rows}
    hotspot_seq15_set = {r["seq15"] for r in hotspot_rows}

    for r in hotspot_rows:
        gseq = gene_seq[r["gene"]]
        centers: List[int] = []
        if r["seq15"] and len(r["seq15"]) == 15:
            for s in find_all_matches(gseq, r["seq15"]):
                centers.append(s + 7)
        if not centers and r["seq9"] and len(r["seq9"]) == 9:
            for s in find_all_matches(gseq, r["seq9"]):
                centers.append(s + 4)
        for c in centers:
            hotspot_centers_by_gene[r["gene"]].append(c)
            hs_start, hs_end = max(0, c - 7), min(len(gseq) - 1, c + 7)
            hotspot_intervals_by_gene[r["gene"]].append((hs_start, hs_end))

    # Build candidates per gene
    candidates_by_gene: Dict[str, List[int]] = defaultdict(list)
    for g in genes:
        gseq = gene_seq[g]
        centers = hotspot_centers_by_gene.get(g, [])
        protected = hotspot_intervals_by_gene.get(g, [])
        if not centers:
            continue
        for hc in centers:
            low = max(7, hc - args.nearby_max_distance)
            high = min(len(gseq) - 8, hc + args.nearby_max_distance)
            for c in range(low, high + 1):
                c15_start, c15_end = c - 7, c + 7
                c9_start, c9_end = c - 4, c + 4
                blocked = False
                for hs, he in protected:
                    if intervals_overlap(c15_start, c15_end, hs, he, args.min_gap):
                        blocked = True
                        break
                if blocked:
                    continue
                seq15 = gseq[c15_start : c15_end + 1]
                seq9 = gseq[c9_start : c9_end + 1]
                if len(seq15) != 15 or len(seq9) != 9:
                    continue
                if seq9 in hotspot_seq9_set or seq15 in hotspot_seq15_set:
                    continue
                candidates_by_gene[g].append(c)

    # Sample controls
    controls: List[Dict[str, object]] = []
    gene_list = [g for g in genes if candidates_by_gene.get(g)]
    if not gene_list:
        raise RuntimeError("No control candidates found.")

    while len(controls) < args.target_controls:
        g = rng.choice(gene_list)
        pool = candidates_by_gene[g]
        if not pool:
            gene_list = [x for x in gene_list if candidates_by_gene.get(x)]
            if not gene_list:
                break
            continue
        c = pool.pop(rng.randrange(len(pool)))
        gseq = gene_seq[g]
        seq15 = gseq[c - 7 : c + 8]
        seq9 = gseq[c - 4 : c + 5]
        controls.append({"gene": g, "seq15": seq15, "seq9": seq9})

    # MGW/ProT for hotspots and controls
    cmd = find_deepdnashape_cmd()
    predictor_obj = None
    if not cmd:
        predictor_obj = load_deepdnashape_predictor()
    if not cmd and predictor_obj is None:
        raise RuntimeError("deepDNAshape not available in this environment.")

    control_seq9 = [c["seq9"] for c in controls]
    hotspot_seq9 = [r["seq9"] for r in hotspot_rows]
    if cmd:
        mgw_ctl = predict_feature_averages(cmd, control_seq9, "MGW", args.layer)
        prot_ctl = predict_feature_averages(cmd, control_seq9, "ProT", args.layer)
        mgw_hot = predict_feature_averages(cmd, hotspot_seq9, "MGW", args.layer)
        prot_hot = predict_feature_averages(cmd, hotspot_seq9, "ProT", args.layer)
    else:
        mgw_ctl = predict_feature_averages_api(predictor_obj, control_seq9, "MGW", args.layer)
        prot_ctl = predict_feature_averages_api(predictor_obj, control_seq9, "ProT", args.layer)
        mgw_hot = predict_feature_averages_api(predictor_obj, hotspot_seq9, "MGW", args.layer)
        prot_hot = predict_feature_averages_api(predictor_obj, hotspot_seq9, "ProT", args.layer)

    # Write workbook with Hotspots + Controls
    out_wb = Workbook()
    ws_hot = out_wb.active
    ws_hot.title = "Hotspots"

    # Copy input hotspot rows and overwrite MGW/ProT columns
    hdr = header[:]
    if "Average_MGW" not in hdr:
        hdr.append("Average_MGW")
    if "Average_ProT" not in hdr:
        hdr.append("Average_ProT")
    ws_hot.append(hdr)

    for i, r in enumerate(hotspot_rows):
        # Rebuild row from original worksheet to preserve columns
        row_idx = i + 2  # since hotspot_rows are in the same order as rows read
        row_vals = []
        for c in range(1, ws.max_column + 1):
            row_vals.append(ws.cell(row_idx, c).value)
        # Extend if new columns were added
        while len(row_vals) < len(hdr):
            row_vals.append(None)
        # Set recalculated MGW/ProT at the correct positions
        mgw_col = hdr.index("Average_MGW")
        prot_col = hdr.index("Average_ProT")
        row_vals[mgw_col] = mgw_hot[i] if i < len(mgw_hot) else None
        row_vals[prot_col] = prot_hot[i] if i < len(prot_hot) else None
        ws_hot.append(row_vals)

    ws_out = out_wb.create_sheet("Controls")
    ws_out.append(["Gene", "Sequence_15_Long", "Sequence_9_Long", "Average_MGW", "Average_ProT"])
    for i, c in enumerate(controls):
        ws_out.append(
            [
                c["gene"],
                c["seq15"],
                c["seq9"],
                mgw_ctl[i] if i < len(mgw_ctl) else None,
                prot_ctl[i] if i < len(prot_ctl) else None,
            ]
        )
    out_wb.save(args.output_xlsx)
    print(f"Controls written: {len(controls)}")
    print(f"Output: {args.output_xlsx}")


if __name__ == "__main__":
    main()
