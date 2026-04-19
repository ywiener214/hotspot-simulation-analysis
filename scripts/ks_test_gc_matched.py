#!/usr/bin/env python3
"""
KS test for MGW and ProT using GC-matched controls (5,000).

Inputs:
- Workbook with Hotspots sheet (with Average_MGW / Average_ProT and Sequence_9_Long)
- Controls workbook or same workbook with Controls sheet

Matching:
- GC_count = number of G/C in Sequence_9_Long (0..9)
- CG_count = number of "CG" dinucleotides in Sequence_9_Long (0..8)
- Controls are matched to hotspots by exact GC_count and optional exact CG_count
  (and same gene if present)

Outputs:
- Prints KS statistics for MGW and ProT
- Prints effect sizes:
  - mean difference with bootstrap CI
  - median difference with bootstrap CI
"""

from __future__ import annotations

import argparse
import math
import random
import statistics
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


DNA = set("ACGTN")
SEQ9_COL = "Sequence_9_Long"
GENE_COL = "Gene"
MGW_COL = "Average_MGW"
PROT_COL = "Average_ProT"


def clean_seq(value: object) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    s = "".join(ch for ch in s if ch in DNA)
    return s or None


def gc_count(seq: str) -> int:
    return sum(1 for ch in seq if ch in ("G", "C"))


def cg_count(seq: str) -> int:
    return sum(1 for i in range(len(seq) - 1) if seq[i : i + 2] == "CG")


def ks_test_2samp(x: List[float], y: List[float]) -> Tuple[float, float]:
    x = sorted([v for v in x if v is not None])
    y = sorted([v for v in y if v is not None])
    n1, n2 = len(x), len(y)
    if n1 == 0 or n2 == 0:
        return float("nan"), float("nan")
    i = j = 0
    d = 0.0
    while i < n1 and j < n2:
        if x[i] <= y[j]:
            i += 1
        else:
            j += 1
        d = max(d, abs(i / n1 - j / n2))
    en = math.sqrt(n1 * n2 / (n1 + n2))
    lam = (en + 0.12 + 0.11 / en) * d
    p = 2.0 * sum(((-1) ** (k - 1)) * math.exp(-2 * (lam ** 2) * (k ** 2)) for k in range(1, 100))
    p = max(0.0, min(1.0, p))
    return d, p


def percentile(sorted_values: List[float], p: float) -> float:
    if not sorted_values:
        return float("nan")
    if len(sorted_values) == 1:
        return sorted_values[0]
    idx = (len(sorted_values) - 1) * p
    lo = math.floor(idx)
    hi = math.ceil(idx)
    if lo == hi:
        return sorted_values[lo]
    frac = idx - lo
    return sorted_values[lo] * (1 - frac) + sorted_values[hi] * frac


def bootstrap_diff_ci(
    x: List[float],
    y: List[float],
    stat_fn,
    rng: random.Random,
    n_boot: int,
    ci: float,
) -> Tuple[float, float, float]:
    x_clean = [float(v) for v in x if v is not None]
    y_clean = [float(v) for v in y if v is not None]
    if not x_clean or not y_clean:
        return float("nan"), float("nan"), float("nan")
    observed = stat_fn(x_clean) - stat_fn(y_clean)
    boot: List[float] = []
    for _ in range(n_boot):
        sx = [x_clean[rng.randrange(len(x_clean))] for _ in range(len(x_clean))]
        sy = [y_clean[rng.randrange(len(y_clean))] for _ in range(len(y_clean))]
        boot.append(stat_fn(sx) - stat_fn(sy))
    boot.sort()
    alpha = (1.0 - ci) / 2.0
    lower = percentile(boot, alpha)
    upper = percentile(boot, 1.0 - alpha)
    return observed, lower, upper


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--hotspots-xlsx", required=True, help="Workbook with Hotspots sheet")
    p.add_argument("--controls-xlsx", required=True, help="Workbook with Controls sheet")
    p.add_argument("--hotspots-sheet", default="Hotspots")
    p.add_argument("--controls-sheet", default="Controls")
    p.add_argument("--target-controls", type=int, default=5000)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--match-gene", action="store_true", help="Also require same-gene when matching")
    p.add_argument(
        "--match-cg-density",
        action="store_true",
        help='Also require exact CG dinucleotide count in the 9-mer when matching',
    )
    p.add_argument("--bootstrap-reps", type=int, default=5000, help="Bootstrap repetitions for CI estimation")
    p.add_argument("--ci", type=float, default=0.95, help="Confidence level for effect-size intervals")
    return p.parse_args()


def load_rows(path: str, sheet: str) -> List[Dict[str, object]]:
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {path}")
    ws = wb[sheet]
    header = [cell.value for cell in ws[1]]
    col_map: Dict[str, int] = {}
    for idx, name in enumerate(header, start=1):
        if isinstance(name, str) and name.strip():
            col_map[name.strip()] = idx
    for required in (SEQ9_COL, MGW_COL, PROT_COL):
        if required not in col_map:
            raise ValueError(f"Missing column {required} in {path}:{sheet}")

    rows: List[Dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        seq9 = clean_seq(ws.cell(r, col_map[SEQ9_COL]).value)
        if not seq9 or len(seq9) != 9:
            continue
        row = {
            "gene": str(ws.cell(r, col_map.get(GENE_COL, 0)).value).strip() if GENE_COL in col_map else None,
            "seq9": seq9,
            "gc": gc_count(seq9),
            "cg": cg_count(seq9),
            "mgw": ws.cell(r, col_map[MGW_COL]).value,
            "prot": ws.cell(r, col_map[PROT_COL]).value,
        }
        rows.append(row)
    return rows


def main() -> None:
    args = parse_args()
    rng = random.Random(args.seed)

    hotspots = load_rows(args.hotspots_xlsx, args.hotspots_sheet)
    controls = load_rows(args.controls_xlsx, args.controls_sheet)

    # Build control pools by exact matching key.
    pool: Dict[Tuple[int, Optional[int], Optional[str]], List[Dict[str, object]]] = defaultdict(list)
    for c in controls:
        key = (
            c["gc"],
            c["cg"] if args.match_cg_density else None,
            c["gene"] if args.match_gene else None,
        )
        pool[key].append(c)

    matched_hotspots: List[Dict[str, object]] = []
    matched_controls: List[Dict[str, object]] = []
    for h in hotspots:
        key = (
            h["gc"],
            h["cg"] if args.match_cg_density else None,
            h["gene"] if args.match_gene else None,
        )
        candidates = pool.get(key, [])
        if not candidates:
            continue
        idx = rng.randrange(len(candidates))
        matched_hotspots.append(h)
        matched_controls.append(candidates.pop(idx))
        if len(matched_controls) >= args.target_controls:
            break

    mgw_hot = [h["mgw"] for h in matched_hotspots if h["mgw"] is not None]
    prot_hot = [h["prot"] for h in matched_hotspots if h["prot"] is not None]
    mgw_ctl = [c["mgw"] for c in matched_controls if c["mgw"] is not None]
    prot_ctl = [c["prot"] for c in matched_controls if c["prot"] is not None]

    d_mgw, p_mgw = ks_test_2samp(mgw_hot, mgw_ctl)
    d_prot, p_prot = ks_test_2samp(prot_hot, prot_ctl)
    mean_mgw, mean_mgw_lo, mean_mgw_hi = bootstrap_diff_ci(
        mgw_hot, mgw_ctl, statistics.mean, rng, args.bootstrap_reps, args.ci
    )
    mean_prot, mean_prot_lo, mean_prot_hi = bootstrap_diff_ci(
        prot_hot, prot_ctl, statistics.mean, rng, args.bootstrap_reps, args.ci
    )
    med_mgw, med_mgw_lo, med_mgw_hi = bootstrap_diff_ci(
        mgw_hot, mgw_ctl, statistics.median, rng, args.bootstrap_reps, args.ci
    )
    med_prot, med_prot_lo, med_prot_hi = bootstrap_diff_ci(
        prot_hot, prot_ctl, statistics.median, rng, args.bootstrap_reps, args.ci
    )

    print(f"Hotspots matched: {len(matched_hotspots)}")
    print(f"Controls matched: {len(matched_controls)} (target {args.target_controls})")
    print(
        "Matching: exact GC_count; "
        + ("exact CG_count; " if args.match_cg_density else "")
        + "gene matching: "
        + ("ON" if args.match_gene else "OFF")
    )
    print(
        'Definitions: GC_count = count(G/C) in 9-mer; '
        'CG_count = count of "CG" dinucleotides in 9-mer.'
    )
    print(f"KS test MGW: D={d_mgw:.4f}, p={p_mgw:.4g}")
    print(f"KS test ProT: D={d_prot:.4f}, p={p_prot:.4g}")
    ci_pct = int(round(args.ci * 100))
    print(
        f"MGW mean difference (hotspots-controls): {mean_mgw:.6f} "
        f"[{ci_pct}% CI {mean_mgw_lo:.6f}, {mean_mgw_hi:.6f}]"
    )
    print(
        f"MGW median difference (hotspots-controls): {med_mgw:.6f} "
        f"[{ci_pct}% CI {med_mgw_lo:.6f}, {med_mgw_hi:.6f}]"
    )
    print(
        f"ProT mean difference (hotspots-controls): {mean_prot:.6f} "
        f"[{ci_pct}% CI {mean_prot_lo:.6f}, {mean_prot_hi:.6f}]"
    )
    print(
        f"ProT median difference (hotspots-controls): {med_prot:.6f} "
        f"[{ci_pct}% CI {med_prot_lo:.6f}, {med_prot_hi:.6f}]"
    )


if __name__ == "__main__":
    main()
